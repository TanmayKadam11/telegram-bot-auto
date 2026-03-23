from telethon import TelegramClient
from openpyxl import Workbook, load_workbook
from datetime import datetime
import asyncio
import os
import re

async def wait_next_minute():
    now = datetime.now()
    seconds = now.second
    wait = 60 - seconds
    print(f"⏳ Waiting {wait} sec for next minute...")
    await asyncio.sleep(wait)

# 🔑 YOUR API
api_id = 23719051
api_hash = "83f2d45bf14e8efac72bbb71f28ffc56"

# 🤖 YOUR BOTS
bots = ['OkWinSureShot_bot', 'wingojalwahack_bot']

# 📁 Excel file
file_name = "results.xlsx"

# 📡 Client
client = TelegramClient('session_vps', api_id, api_hash)

# 📊 Excel setup
if not os.path.exists(file_name):
    wb = Workbook()
    ws = wb.active
    ws.append(["Time", "Bot", "Period", "Purchase"])
    wb.save(file_name)

# 🔍 Extract data
def extract_data(text):
    period = None
    purchase = None

    lines = text.split('\n')

    for line in lines:
        if "Period" in line:
            p = re.search(r'\d+', line)
            if p:
                period = p.group()

        if "Purchase" in line:
            if "Big" in line:
                purchase = "Big"
            elif "Small" in line:
                purchase = "Small"

    return period, purchase

# 💾 Save to Excel
def save_to_excel(bot, period, purchase):
    wb = load_workbook(file_name)
    ws = wb.active

    ws.append([
        datetime.now().strftime("%H:%M:%S"),
        bot,
        period,
        purchase
    ])

    wb.save(file_name)

# 🔁 MAIN LOOP
async def main():
    await client.start()
    print("🚀 Script Started...\n")

    await wait_next_minute()

    while True:
        for bot in bots:
            try:
                print(f"👉 Processing {bot}")

                # Step 1: Send message (ONLY ONCE)
                await client.send_message(bot, "🔮 Get Prediction")
                print("📤 Sent: Get Prediction")

                await asyncio.sleep(3)

                # Step 2: Get messages
                msgs = await client.get_messages(bot, limit=10)

                clicked = False

                # Step 3: Try button click (if exists)
                for msg in reversed(msgs):
                    if msg.buttons:
                        print("🟢 Clicking button")
                        await msg.click(0)
                        clicked = True
                        break

                # Step 4: अगर button नहीं मिला तो ignore
                if not clicked:
                    print("⚠ Button not found (text mode working)")

                await asyncio.sleep(5)

                # 📥 Read messages
                msgs = await client.get_messages(bot, limit=10)

                found = False

                for msg in reversed(msgs):
                    text = msg.text

                    if text and "Period" in text:
                        data = extract_data(text)

                        if data and data[0] and data[1]:
                            period, purchase = data
                            print(f"✅ {bot} -> {period} | {purchase}")

                            save_to_excel(bot, period, purchase)
                            found = True
                            break

                if not found:
                    print("⚠ Data not found")

            except Exception as e:
                print(f"❌ Error in {bot}: {e}")

        print("⏳ Waiting 60 seconds...\n")
        await wait_next_minute()

# ▶ RUN
with client:
    client.loop.run_until_complete(main())